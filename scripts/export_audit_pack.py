from __future__ import annotations

import argparse
import json
import subprocess
import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES = ROOT / "templates"
OUTPUT = ROOT / "output"


def load_payload(path: Path) -> dict:
    if path.suffix.lower() == ".js":
        return load_js_payload(path)
    return json.loads(path.read_text(encoding="utf8"))


def load_js_payload(path: Path) -> dict:
    node_script = """
const fs = require("fs");
const vm = require("vm");
const filePath = process.argv[1];
const code = fs.readFileSync(filePath, "utf8");
const context = { window: {} };
vm.createContext(context);
vm.runInContext(code, context);
process.stdout.write(JSON.stringify(context.window.SHIP_SENTINEL_DEMO_STATE));
""".strip()
    result = subprocess.run(
        ["node", "-e", node_script, str(path)],
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(result.stdout)


def normalize_payload(payload: dict) -> dict:
    if "executions" in payload and "meta" in payload and payload.get("memo", {}).get("scope_covered") is not None:
        return payload
    return normalize_app_state(payload)


def normalize_app_state(payload: dict) -> dict:
    meta = payload.get("meta", {})
    intake = payload.get("intake", {})
    screens = {item.get("screen_id"): item for item in payload.get("screens", [])}
    bugs = payload.get("bugs", [])
    evidence = payload.get("evidence", [])
    notes = payload.get("notes", [])
    memo = payload.get("memo", {})

    normalized_scenarios = []
    normalized_executions = []

    for index, scenario in enumerate(payload.get("scenarios", []), start=1):
        screen = screens.get(scenario.get("screen_id"), {})
        note = first_note_for(notes, scenario.get("screen_id"), scenario.get("scenario_id"))
        linked_evidence = first_evidence_for(evidence, scenario.get("scenario_id"), scenario.get("screen_id"))
        linked_bug = first_bug_for(bugs, scenario.get("scenario_id"), scenario.get("screen_id"))
        normalized_scenarios.append(
            {
                "scenario_id": scenario.get("scenario_id", f"SCN-{index:03d}"),
                "feature_area": screen.get("feature_area", "Unassigned"),
                "priority": scenario.get("priority", "P1"),
                "title": scenario.get("title", "Untitled scenario"),
                "user_goal": scenario.get("title", "Scenario execution"),
                "preconditions": build_preconditions(intake, scenario),
                "steps": scenario.get("steps", ""),
                "expected_result": scenario.get("expected_result", ""),
                "risk_tag": screen.get("feature_area", "General"),
                "status": scenario.get("status", "미실행"),
                "owner_notes": note.get("body", ""),
            }
        )

        if scenario.get("status") and scenario.get("status") != "미실행":
            normalized_executions.append(
                {
                    "execution_id": f"EXE-{len(normalized_executions) + 1:03d}",
                    "scenario_id": scenario.get("scenario_id", f"SCN-{index:03d}"),
                    "executed_at": linked_evidence.get("captured_at", meta.get("delivery_date", "")),
                    "tester": meta.get("requested_by", "QA Operator"),
                    "environment": meta.get("environment", "Staging"),
                    "browser_device": ", ".join(intake.get("browsers", [])) or "Unknown",
                    "result": scenario.get("status", ""),
                    "observed_result": linked_bug.get("actual_result") or linked_evidence.get("summary", scenario.get("status", "")),
                    "evidence_id": linked_evidence.get("evidence_id", ""),
                    "linked_bug_id": linked_bug.get("bug_id", ""),
                    "retest_required": "Yes" if scenario.get("status") in {"실패", "재검수"} else "No",
                    "notes": note.get("body", ""),
                }
            )

    normalized_bugs = []
    for bug in bugs:
        scenario = first_by_id(payload.get("scenarios", []), "scenario_id", bug.get("scenario_id"))
        screen = screens.get(bug.get("screen_id"), {})
        linked_evidence = first_evidence_for(evidence, bug.get("scenario_id"), bug.get("screen_id"))
        normalized_bugs.append(
            {
                "bug_id": bug.get("bug_id", ""),
                "severity": bug.get("severity", "중간"),
                "category": bug.get("drift_type", "결함"),
                "title": bug.get("title", ""),
                "environment": meta.get("environment", "Staging"),
                "browser_device": ", ".join(intake.get("browsers", [])) or "Unknown",
                "scenario_id": bug.get("scenario_id", ""),
                "repro_steps": scenario.get("steps", ""),
                "expected_result": bug.get("expected_result") or scenario.get("expected_result", ""),
                "actual_result": bug.get("actual_result") or linked_evidence.get("summary", ""),
                "impact": bug.get("impact", ""),
                "evidence_id": linked_evidence.get("evidence_id", ""),
                "owner_recommendation": meta.get("recommended_next_owner", ""),
                "status": bug.get("status", "열림"),
            }
        )

    normalized_evidence = []
    for item in evidence:
        linked_bug = first_bug_for(bugs, item.get("scenario_id"), item.get("screen_id"))
        normalized_evidence.append(
            {
                "evidence_id": item.get("evidence_id", ""),
                "type": "스크린샷",
                "captured_at": item.get("captured_at", ""),
                "file_name": item.get("file_name", f"{item.get('evidence_id', 'evidence')}.png"),
                "scenario_id": item.get("scenario_id", ""),
                "bug_id": linked_bug.get("bug_id", ""),
                "source_url": build_source_url(meta.get("target_url", ""), screens.get(item.get("screen_id"), {}).get("route_hint", "")),
                "summary": item.get("summary", ""),
                "storage_path": f"evidence/{item.get('file_name', item.get('evidence_id', 'capture.png'))}",
                "reviewer_notes": "" if not item.get("is_reference_candidate") else "레퍼런스 후보 캡처",
            }
        )

    normalized_memo = {
        "decision": memo.get("decision", "검수 진행중"),
        "headline": memo.get("headline", "릴리즈 메모 준비 중"),
        "scope_covered": build_scope_covered(payload),
        "blockers": ensure_list(memo.get("blockers")),
        "warnings": ensure_list(memo.get("warnings")),
        "passed_areas": build_passed_areas(payload),
        "actions": ensure_list(memo.get("actions")),
    }

    return {
        "meta": {
            "audit_id": meta.get("audit_id", "audit-unknown"),
            "service_name": meta.get("service_name", "Unknown Service"),
            "target_url": meta.get("target_url", ""),
            "release_goal": meta.get("release_goal", ""),
            "requested_by": meta.get("requested_by", "QA Operator"),
            "delivery_date": meta.get("delivery_date", ""),
            "recommended_next_owner": meta.get("recommended_next_owner", "QA Lead"),
            "package_version": meta.get("package_version", "v0.1"),
        },
        "intake": intake,
        "scenarios": normalized_scenarios,
        "executions": normalized_executions,
        "bugs": normalized_bugs,
        "evidence": normalized_evidence,
        "memo": normalized_memo,
    }


def build_preconditions(intake: dict, scenario: dict) -> str:
    accounts = intake.get("accounts", "")
    parts = [item for item in [accounts, scenario.get("preconditions")] if item]
    return " / ".join(parts)


def build_source_url(target_url: str, route_hint: str) -> str:
    return f"{target_url}{route_hint}" if route_hint else target_url


def build_scope_covered(payload: dict) -> str:
    intake = payload.get("intake", {})
    flows = ensure_list(intake.get("critical_flows"))
    if flows:
        return ", ".join(flows)
    screen_names = [item.get("display_name", item.get("screen_id", "")) for item in payload.get("screens", [])]
    return ", ".join(item for item in screen_names if item) or "핵심 범위 미정"


def build_passed_areas(payload: dict) -> list[str]:
    screens = {item.get("screen_id"): item for item in payload.get("screens", [])}
    passed = []
    for scenario in payload.get("scenarios", []):
        if scenario.get("status") != "통과":
            continue
        screen = screens.get(scenario.get("screen_id"), {})
        label = screen.get("display_name") or scenario.get("title")
        if label and label not in passed:
            passed.append(label)
    return passed or ["아직 통과 영역 없음"]


def first_by_id(items: list[dict], key: str, value: str | None) -> dict:
    for item in items:
        if item.get(key) == value:
            return item
    return {}


def first_bug_for(items: list[dict], scenario_id: str | None, screen_id: str | None) -> dict:
    for item in items:
        if scenario_id and item.get("scenario_id") == scenario_id:
            return item
    for item in items:
        if screen_id and item.get("screen_id") == screen_id:
            return item
    return {}


def first_evidence_for(items: list[dict], scenario_id: str | None, screen_id: str | None) -> dict:
    for item in items:
        if scenario_id and item.get("scenario_id") == scenario_id:
            return item
    for item in items:
        if screen_id and item.get("screen_id") == screen_id:
            return item
    return {}


def first_note_for(items: list[dict], screen_id: str | None, scenario_id: str | None) -> dict:
    for item in items:
        if scenario_id and item.get("scenario_id") == scenario_id:
            return item
    for item in items:
        if screen_id and item.get("screen_id") == screen_id:
            return item
    return {}


def ensure_list(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if value is None:
        return []
    text = str(value).strip()
    return [text] if text else []


def write_rows(workbook_path: Path, sheet_name: str, rows: list[list[object]], start_row: int = 2) -> None:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    max_data_row = ws.max_row
    if max_data_row >= start_row:
        ws.delete_rows(start_row, max_data_row - start_row + 1)
    for row in rows:
        ws.append(row)
    wb.save(workbook_path)


def export_workbooks(payload: dict, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    mappings = [
        ("test-scenarios-template.xlsx", "test-scenarios.xlsx", "Scenarios", [[
            item["scenario_id"], item["feature_area"], item["priority"], item["title"], item["user_goal"],
            item["preconditions"], item["steps"], item["expected_result"], item["risk_tag"], item["status"], item["owner_notes"],
        ] for item in payload.get("scenarios", [])]),
        ("execution-results-template.xlsx", "execution-results.xlsx", "Execution Results", [[
            item["execution_id"], item["scenario_id"], item["executed_at"], item["tester"], item["environment"],
            item["browser_device"], item["result"], item["observed_result"], item["evidence_id"], item["linked_bug_id"],
            item["retest_required"], item["notes"],
        ] for item in payload.get("executions", [])]),
        ("bug-register-template.xlsx", "bug-register.xlsx", "Bug Register", [[
            item["bug_id"], item["severity"], item["category"], item["title"], item["environment"], item["browser_device"],
            item["scenario_id"], item["repro_steps"], item["expected_result"], item["actual_result"], item["impact"],
            item["evidence_id"], item["owner_recommendation"], item["status"],
        ] for item in payload.get("bugs", [])]),
        ("evidence-index-template.xlsx", "evidence-index.xlsx", "Evidence Index", [[
            item["evidence_id"], item["type"], item["captured_at"], item["file_name"], item["scenario_id"], item["bug_id"],
            item["source_url"], item["summary"], item["storage_path"], item["reviewer_notes"],
        ] for item in payload.get("evidence", [])]),
        ("release-memo-inputs-template.xlsx", "release-memo-inputs.xlsx", "Release Memo Inputs", [
            ["Scope covered", payload["memo"]["scope_covered"], "QA Lead", "P0", "Context"],
            ["Blockers", "\n".join(payload["memo"]["blockers"]), payload["meta"]["recommended_next_owner"], "P0", "Block"],
            ["Warnings", "\n".join(payload["memo"]["warnings"]), "Frontend Lead", "P1", "Fix before ship"],
            ["Passed areas", "\n".join(payload["memo"]["passed_areas"]), "QA Lead", "P1", "Confidence"],
            ["Recommendation", payload["memo"]["headline"], payload["meta"]["recommended_next_owner"], "P0", payload["memo"]["decision"]],
        ]),
    ]

    for template_name, output_name, sheet_name, rows in mappings:
        source = TEMPLATES / template_name
        target = output_dir / output_name
        target.write_bytes(source.read_bytes())
        write_rows(target, sheet_name, rows)


def export_release_memo(payload: dict, output_dir: Path) -> None:
    memo_path = output_dir / "release-memo.md"
    lines = [
        f"# Release Memo: {payload['meta']['service_name']}",
        "",
        f"- Audit ID: `{payload['meta']['audit_id']}`",
        f"- Target: `{payload['meta']['target_url']}`",
        f"- Requested by: `{payload['meta']['requested_by']}`",
        f"- Delivery date: `{payload['meta']['delivery_date']}`",
        f"- Decision: `{payload['memo']['decision']}`",
        f"- Recommended next owner: `{payload['meta']['recommended_next_owner']}`",
        "",
        "## Scope Covered",
        payload["memo"]["scope_covered"],
        "",
        "## Blockers",
    ]
    lines.extend(f"- {item}" for item in payload["memo"]["blockers"])
    lines.extend(["", "## Warnings"])
    lines.extend(f"- {item}" for item in payload["memo"]["warnings"])
    lines.extend(["", "## Passed Areas"])
    lines.extend(f"- {item}" for item in payload["memo"]["passed_areas"])
    lines.extend(["", "## Actions"])
    lines.extend(f"- {item}" for item in payload["memo"]["actions"])
    memo_path.write_text("\n".join(lines), encoding="utf8")


def export_supporting_files(payload: dict, output_dir: Path) -> None:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    evidence_manifest = {
        "audit_id": payload["meta"]["audit_id"],
        "package_version": payload["meta"].get("package_version", "v0.1"),
        "generated_at": generated_at,
        "items": payload.get("evidence", []),
    }
    (output_dir / "evidence-bundle-manifest.json").write_text(
        json.dumps(evidence_manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf8",
    )

    generated_files = sorted(path.name for path in output_dir.iterdir() if path.is_file())
    package_metadata = {
        "audit_id": payload["meta"]["audit_id"],
        "service_name": payload["meta"]["service_name"],
        "package_version": payload["meta"].get("package_version", "v0.1"),
        "generated_at": generated_at,
        "delivery_slug": f"{payload['meta']['audit_id']}__{payload['meta'].get('package_version', 'v0.1')}",
        "generated_files": generated_files,
        "client_bundle_files": [
            "client-summary.md",
            "package-diff.md",
            "delivery-history.md",
            "release-memo.md"
        ],
    }
    (output_dir / "package-metadata.json").write_text(
        json.dumps(package_metadata, ensure_ascii=False, indent=2) + "\n",
        encoding="utf8",
    )

    # Persist normalized audit payload for diffing
    (output_dir / "normalized-audit.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf8",
    )



def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def write_supporting_files_and_metadata(payload: dict, output_dir: Path) -> dict:
    """Write evidence manifest and package metadata with file hashes.

    Returns the written metadata dict.
    """
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    evidence_manifest = {
        "audit_id": payload["meta"]["audit_id"],
        "package_version": payload["meta"].get("package_version", "v0.1"),
        "generated_at": generated_at,
        "items": payload.get("evidence", []),
    }
    (output_dir / "evidence-bundle-manifest.json").write_text(
        json.dumps(evidence_manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf8",
    )

    generated_files = sorted(path.name for path in output_dir.iterdir() if path.is_file())
    file_hashes = {name: _sha256(output_dir / name) for name in generated_files}
    package_metadata = {
        "audit_id": payload["meta"]["audit_id"],
        "service_name": payload["meta"]["service_name"],
        "package_version": payload["meta"].get("package_version", "v0.1"),
        "generated_at": generated_at,
        "delivery_slug": f"{payload['meta']['audit_id']}__{payload['meta'].get('package_version', 'v0.1')}",
        "generated_files": generated_files,
        "file_hashes": file_hashes,
    }
    (output_dir / "package-metadata.json").write_text(
        json.dumps(package_metadata, ensure_ascii=False, indent=2) + "\n",
        encoding="utf8",
    )
    return package_metadata


def _list_to_set_str(items: list[str] | None) -> set[str]:
    return {str(x) for x in (items or []) if str(x).strip()}


def summarize_counts(payload: dict) -> dict:
    return {
        'scenarios': len(payload.get('scenarios', [])),
        'executions': len(payload.get('executions', [])),
        'bugs': len(payload.get('bugs', [])),
        'evidence': len(payload.get('evidence', [])),
        'passed': sum(1 for s in payload.get('scenarios', []) if s.get('status') == '통과'),
        'failed': sum(1 for s in payload.get('scenarios', []) if s.get('status') in {'실패', '재검수'}),
    }


def diff_payloads(prev: dict | None, curr: dict) -> dict:
    prev = prev or {}
    prev_scen = {s.get('scenario_id'): s for s in prev.get('scenarios', [])}
    curr_scen = {s.get('scenario_id'): s for s in curr.get('scenarios', [])}

    added_scenarios = sorted([sid for sid in curr_scen.keys() if sid not in prev_scen])
    removed_scenarios = sorted([sid for sid in prev_scen.keys() if sid not in curr_scen])
    status_changes = []
    for sid in curr_scen.keys() & prev_scen.keys():
        a = prev_scen[sid].get('status')
        b = curr_scen[sid].get('status')
        if a != b:
            status_changes.append({
                'scenario_id': sid,
                'title': curr_scen[sid].get('title') or prev_scen[sid].get('title'),
                'from': a,
                'to': b,
            })
    status_changes.sort(key=lambda x: x['scenario_id'])

def write_package_diff_md(diff: dict, output_dir: Path) -> None:
    lines = []
    lines.append(f"# Package Diff — {diff.get('audit_id','-')}")
    lines.append("")
    lines.append(f"- Current generated_at: `{diff.get('generated_at','-')}`")
    pg = diff.get('previous_generated_at')
    if pg:
        lines.append(f"- Previous generated_at: `{pg}`")
    lines.append("")
    pc = diff.get('payload_changes') or {}
    fc = diff.get('file_changes') or {}
    from_c = pc.get('from_counts', {})
    to_c = pc.get('to_counts', {})
    def row(label):
        a = int(from_c.get(label, 0) or 0)
        b = int(to_c.get(label, 0) or 0)
        return f"- {label}: {a} → {b} (Δ {b-a:+d})"
    for label in ('scenarios','executions','bugs','evidence','passed','failed'):
        lines.append(row(label))
    lines.append("")
    sc = pc.get('scenario_status_changes') or []
    if sc:
        lines.append('## Scenario status changes')
        for item in sc:
            lines.append(f"- {item.get('scenario_id')}: {item.get('from') or '-'} → {item.get('to') or '-'} — {item.get('title','')}")
        lines.append("")
    if pc.get('added_bug_ids') or pc.get('resolved_bug_ids'):
        lines.append('## Bugs')
        if pc.get('added_bug_ids'):
            lines.append('- Added: ' + ', '.join(f"`{x}`" for x in pc['added_bug_ids']))
        if pc.get('resolved_bug_ids'):
            lines.append('- Resolved: ' + ', '.join(f"`{x}`" for x in pc['resolved_bug_ids']))
        lines.append("")
    if any(fc.get(k) for k in ('added','removed','changed')):
        lines.append('## Files')
        for k in ('added','removed','changed'):
            vals = fc.get(k) or []
            if vals:
                lines.append(f"- {k.title()}: {len(vals)}")
        lines.append("")



    (output_dir / 'package-diff.md').write_text('\n'.join(lines) + '\n', encoding='utf8')
def index_packages(root_output: Path):
    """Return list of (dir_path, metadata_json) for all output packages that have metadata."""
    results = []
    if not root_output.exists():
        return results
    for child in root_output.iterdir():
        if not child.is_dir():
            continue
        meta_path = child / "package-metadata.json"
        if meta_path.exists():
            try:
                results.append((child, json.loads(meta_path.read_text(encoding="utf8"))))
            except Exception:
                continue
    return results


def _parse_dt(dt: str):
    try:
        return datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.min


def find_previous_package(this_dir: Path, this_meta: dict):
    audit_id = this_meta.get("audit_id")
    if not audit_id:
        return None, None
    candidates = []
    for path, meta in index_packages(OUTPUT):
        if path.resolve() == this_dir.resolve():
            continue
        if meta.get("audit_id") != audit_id:
            continue
        candidates.append((path, meta))
    if not candidates:
        return None, None
    candidates.sort(key=lambda item: _parse_dt(item[1].get("generated_at", "")))
    return candidates[-1]


def _diff_lists_by_id(old_list, new_list, key):
    old_by = {str(item.get(key)): item for item in old_list}
    new_by = {str(item.get(key)): item for item in new_list}
    old_ids = set(old_by.keys())
    new_ids = set(new_by.keys())
    added = sorted(new_ids - old_ids)
    removed = sorted(old_ids - new_ids)
    common = sorted(old_ids & new_ids)
    changed = []
    for cid in common:
        o = old_by[cid]
        n = new_by[cid]
        deltas = {}
        for field in ("status", "severity", "category", "title", "impact"):
            if (o.get(field) != n.get(field)):
                deltas[field] = (o.get(field), n.get(field))
        if deltas:
            changed.append({"id": cid, "changes": deltas})
    return {
        "added": added,
        "removed": removed,
        "changed": changed,
        "old_count": len(old_list),
        "new_count": len(new_list),
        "delta": len(new_list) - len(old_list),
    }


def build_package_diff(prev_dir: Path, prev_meta: dict, curr_dir: Path, curr_meta: dict):
    prev_norm = prev_dir / "normalized-audit.json"
    curr_norm = curr_dir / "normalized-audit.json"
    prev_payload = json.loads(prev_norm.read_text(encoding="utf8")) if prev_norm.exists() else {}
    curr_payload = json.loads(curr_norm.read_text(encoding="utf8")) if curr_norm.exists() else {}

    prev_evidence = prev_payload.get("evidence", [])
    curr_evidence = curr_payload.get("evidence", [])
    if not prev_evidence and (prev_dir / "evidence-bundle-manifest.json").exists():
        prev_evidence = json.loads((prev_dir / "evidence-bundle-manifest.json").read_text(encoding="utf8")).get("items", [])
    if not curr_evidence and (curr_dir / "evidence-bundle-manifest.json").exists():
        curr_evidence = json.loads((curr_dir / "evidence-bundle-manifest.json").read_text(encoding="utf8")).get("items", [])

    prev_scenarios = prev_payload.get("scenarios", [])
    curr_scenarios = curr_payload.get("scenarios", [])
    prev_bugs = prev_payload.get("bugs", [])
    curr_bugs = curr_payload.get("bugs", [])
    prev_executions = prev_payload.get("executions", [])
    curr_executions = curr_payload.get("executions", [])
    prev_notes = prev_payload.get("notes", [])
    curr_notes = curr_payload.get("notes", [])
    prev_changes = prev_payload.get("change_requests", [])
    curr_changes = curr_payload.get("change_requests", [])

    files_old = set((prev_meta or {}).get("generated_files", []))
    files_new = set((curr_meta or {}).get("generated_files", []))

    diff = {
        "meta": {
            "audit_id": curr_meta.get("audit_id") or prev_meta.get("audit_id"),
            "previous_delivery_slug": prev_meta.get("delivery_slug"),
            "current_delivery_slug": curr_meta.get("delivery_slug"),
            "previous_generated_at": prev_meta.get("generated_at"),
            "current_generated_at": curr_meta.get("generated_at"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "files": {
            "added": sorted(list(files_new - files_old)),
            "removed": sorted(list(files_old - files_new)),
            "unchanged": sorted(list(files_new & files_old)),
        },
        "scenarios": _diff_lists_by_id(prev_scenarios, curr_scenarios, "scenario_id"),
        "executions": {
            "old_count": len(prev_executions),
            "new_count": len(curr_executions),
            "delta": len(curr_executions) - len(prev_executions),
        },
        "bugs": _diff_lists_by_id(prev_bugs, curr_bugs, "bug_id"),
        "evidence": _diff_lists_by_id(prev_evidence, curr_evidence, "evidence_id"),
        "notes": _diff_lists_by_id(prev_notes, curr_notes, "note_id"),
        "change_requests": _diff_lists_by_id(prev_changes, curr_changes, "change_request_id"),
        "memo": {
            "decision": {
                "before": (prev_payload.get("memo", {}) or {}).get("decision"),
                "after": (curr_payload.get("memo", {}) or {}).get("decision"),
            },
            "headline_changed": ((prev_payload.get("memo", {}) or {}).get("headline") != (curr_payload.get("memo", {}) or {}).get("headline")),
        },
    }
    return diff


def write_diff_markdown(diff: dict, output_dir: Path):
    lines = []
    m = diff.get("meta", {})
    lines.append(f"# Package Diff: {m.get('previous_delivery_slug', '-') } → {m.get('current_delivery_slug', '-')} ")
    lines.append("")
    lines.append(f"- Audit: `{m.get('audit_id', '-')}`")
    lines.append(f"- Compared at: `{m.get('generated_at', '-')}`")
    lines.append(f"- Previous: `{m.get('previous_delivery_slug', '-')}` @ `{m.get('previous_generated_at', '-')}`")
    lines.append(f"- Current: `{m.get('current_delivery_slug', '-')}` @ `{m.get('current_generated_at', '-')}`")
    lines.append("")
    for key in ("scenarios", "executions", "bugs", "evidence", "notes", "change_requests"):
        section = diff.get(key, {})
        old_c = section.get("old_count", 0)
        new_c = section.get("new_count", 0)
        delta = section.get("delta", new_c - old_c)
        lines.append(f"- {key}: {old_c} → {new_c} (Δ {delta:+d})")
    lines.append("")
    scen = diff.get("scenarios", {})
    changes = scen.get("changed", [])
    if changes:
        lines.append("## Scenario status changes")
        for item in changes:
            cid = item.get("id")
            ch = item.get("changes", {})
            if "status" in ch:
                before, after = ch["status"]
                lines.append(f"- `{cid}`: {before or '-'} → {after or '-'}")
        lines.append("")
    bugs = diff.get("bugs", {})
    if bugs.get("added") or bugs.get("removed"):
        lines.append("## Bugs")
        if bugs.get("added"):
            lines.append("- Added: " + ", ".join(f"`{bid}`" for bid in bugs["added"]))
        if bugs.get("removed"):
            lines.append("- Removed: " + ", ".join(f"`{bid}`" for bid in bugs["removed"]))
        lines.append("")
    ev = diff.get("evidence", {})
    if ev.get("added") or ev.get("removed"):
        lines.append("## Evidence")
        if ev.get("added"):
            lines.append(f"- New evidence: {len(ev['added'])}")
        if ev.get("removed"):
            lines.append(f"- Removed evidence: {len(ev['removed'])}")
        lines.append("")
    (output_dir / "package-diff.md").write_text("\n".join(lines) + "\n", encoding="utf8")


def compute_payload_changes(prev: dict | None, curr: dict) -> dict:
    prev = prev or {}
    def _counts(x: dict) -> dict:
        return {
            'scenarios': len(x.get('scenarios', []) or []),
            'executions': len(x.get('executions', []) or []),
            'bugs': len(x.get('bugs', []) or []),
            'evidence': len(x.get('evidence', []) or []),
            'passed': sum(1 for s in (x.get('scenarios', []) or []) if s.get('status') == '통과'),
            'failed': sum(1 for s in (x.get('scenarios', []) or []) if s.get('status') in {'실패','재검수'}),
        }
    prev_scen = {s.get('scenario_id'): s for s in (prev.get('scenarios', []) or [])}
    curr_scen = {s.get('scenario_id'): s for s in (curr.get('scenarios', []) or [])}
    status_changes = []
    for sid in set(prev_scen.keys()) & set(curr_scen.keys()):
        a = prev_scen[sid].get('status')
        b = curr_scen[sid].get('status')
        if a != b:
            status_changes.append({'scenario_id': sid, 'title': curr_scen[sid].get('title') or prev_scen[sid].get('title'), 'from': a, 'to': b})
    status_changes.sort(key=lambda x: x['scenario_id'])
    prev_bug_ids = [b.get('bug_id') for b in (prev.get('bugs', []) or []) if b.get('bug_id')]
    curr_bug_ids = [b.get('bug_id') for b in (curr.get('bugs', []) or []) if b.get('bug_id')]
    added_bug_ids = sorted([x for x in set(curr_bug_ids) - set(prev_bug_ids)])
    resolved_bug_ids = sorted([x for x in set(prev_bug_ids) - set(curr_bug_ids)])
    return {
        'from_counts': _counts(prev),
        'to_counts': _counts(curr),
        'scenario_status_changes': status_changes,
        'added_bug_ids': added_bug_ids,
        'resolved_bug_ids': resolved_bug_ids,
    }


def _load_normalized_payload(pkg_dir: Path) -> dict:
    path = pkg_dir / 'normalized-payload.json'
    if path.exists():
        return json.loads(path.read_text(encoding='utf8'))
    alt = pkg_dir / 'normalized-audit.json'
    return json.loads(alt.read_text(encoding='utf8')) if alt.exists() else {}


def write_delivery_history(audit_id: str, current_dir: Path, current_meta: dict) -> dict:
    packages = []
    for pkg_path, meta in index_packages(OUTPUT):
        if (meta or {}).get('audit_id') == audit_id:
            packages.append((pkg_path, meta))
    # Python-level check below ensures current is included if missing.
    if not any(p[0].resolve() == current_dir.resolve() for p in packages):
        packages.append((current_dir, current_meta))
    def _dt(meta: dict) -> str:
        return str((meta or {}).get('generated_at') or '')
    packages.sort(key=lambda item: _dt(item[1]))
    history = {'audit_id': audit_id, 'packages': []}
    prev_counts = None
    for pkg_dir, meta in packages:
        payload = _load_normalized_payload(pkg_dir)
        counts = summarize_counts(payload) if payload else {}
        entry = {
            'delivery_slug': meta.get('delivery_slug'),
            'generated_at': meta.get('generated_at'),
            'package_version': meta.get('package_version'),
            'counts': counts,
            'memo': (payload.get('memo', {}) if payload else {}),
        }
        if prev_counts is not None:
            entry['delta'] = {k: int(counts.get(k,0) or 0) - int(prev_counts.get(k,0) or 0) for k in ('scenarios','executions','bugs','evidence','passed','failed')}
        history['packages'].append(entry)
        prev_counts = counts
    (current_dir / 'delivery-history.json').write_text(json.dumps(history, ensure_ascii=False, indent=2) + '\n', encoding='utf8')
    md = []
    md.append(f'# Delivery History — {audit_id}')
    md.append('')
    for i, entry in enumerate(history['packages'], start=1):
        md.append(f"## {i}. {entry.get('delivery_slug','-')} ({entry.get('generated_at','-')})")
        c = entry.get('counts', {})
        md.append(f"- counts: scenarios {c.get('scenarios',0)}, executions {c.get('executions',0)}, bugs {c.get('bugs',0)}, evidence {c.get('evidence',0)}, passed {c.get('passed',0)}, failed {c.get('failed',0)}")
        if entry.get('delta'):
            d = entry['delta']
            md.append(f"- delta: scen {d.get('scenarios',0):+d}, exec {d.get('executions',0):+d}, bugs {d.get('bugs',0):+d}, ev {d.get('evidence',0):+d}, pass {d.get('passed',0):+d}, fail {d.get('failed',0):+d}")
        dec = (entry.get('memo') or {}).get('decision')
        if dec:
            md.append(f"- decision: {dec}")
        md.append('')
    (current_dir / 'delivery-history.md').write_text('\n'.join(md) + '\n', encoding='utf8')
    return history


def write_client_summary(payload: dict, diff: dict | None, output_dir: Path) -> None:
    meta = payload.get('meta', {})
    memo = payload.get('memo', {})
    counts = summarize_counts(payload)
    lines = []
    lines.append(f"# Client Summary — {meta.get('service_name','Service')}")
    lines.append('')
    lines.append(f"- Audit ID: {meta.get('audit_id','-')}")
    lines.append(f"- Package: {meta.get('package_version','v0.1')} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Decision: {memo.get('decision','-')} — {memo.get('headline','')}")
    lines.append(f"- Recommended next owner: {meta.get('recommended_next_owner','-')}")
    lines.append('')
    lines.append('## Current Coverage & Counts')
    lines.append(f"- Scope covered: {memo.get('scope_covered','-')}")
    lines.append(f"- Scenarios {counts.get('scenarios',0)}, Executions {counts.get('executions',0)}, Bugs {counts.get('bugs',0)}, Evidence {counts.get('evidence',0)}, Pass {counts.get('passed',0)}, Fail {counts.get('failed',0)}")
    lines.append('## Audit Narrative')
    summary_bits = []
    summary_bits.append(f"Validated {counts.get('scenarios',0)} scenarios: {counts.get('passed',0)} passed, {counts.get('failed',0)} need attention.")
    if diff:
        pc = diff.get('payload_changes') or {}
        from_c = pc.get('from_counts', {})
        to_c = pc.get('to_counts', {})
        def _d(lbl):
            a = int(from_c.get(lbl, 0) or 0); b = int(to_c.get(lbl, 0) or 0); return b - a
        deltas = [f"Δ scenarios {_d('scenarios'):+d}", f"Δ bugs {_d('bugs'):+d}", f"Δ pass {_d('passed'):+d}", f"Δ fail {_d('failed'):+d}"]
        summary_bits.append('Compared to previous: ' + ', '.join(deltas) + '.')
        if pc.get('added_bug_ids') or pc.get('resolved_bug_ids'):
            newb = ', '.join(pc.get('added_bug_ids', [])[:5]) or 'none'
            resb = ', '.join(pc.get('resolved_bug_ids', [])[:5]) or 'none'
            summary_bits.append(f"New bugs: {newb}. Resolved: {resb}.")
    decision = memo.get('decision')
    if decision:
        summary_bits.append(f"Decision: {decision}.")
    lines.append('- ' + ' '.join(summary_bits))
    lines.append('')
    if diff:
        lines.append('## Changes Since Previous Package')
        pc = diff.get('payload_changes') or {}
        fc = diff.get('file_changes') or {}
        from_c = pc.get('from_counts', {})
        to_c = pc.get('to_counts', {})
        def _delta(label: str) -> str:
            a = int(from_c.get(label, 0) or 0)
            b = int(to_c.get(label, 0) or 0)
            return f"{label}: {a} → {b} (Δ {b-a:+d})"
        lines.append('- ' + ', '.join([_delta(x) for x in ('scenarios','executions','bugs','evidence','passed','failed')]))
        if pc.get('scenario_status_changes'):
            lines.append(f"- Scenario status changes: {len(pc['scenario_status_changes'])}")
        if pc.get('added_bug_ids') or pc.get('resolved_bug_ids'):
            if pc.get('added_bug_ids'):
                lines.append('- New bugs: ' + ', '.join(f"{x}" for x in pc['added_bug_ids']))
            if pc.get('resolved_bug_ids'):
                lines.append('- Resolved bugs: ' + ', '.join(f"{x}" for x in pc['resolved_bug_ids']))
        if any((fc or {}).get(k) for k in ('added','removed','changed')):
            lines.append(f"- Deliverable files changed: added {len(fc.get('added',[]))}, removed {len(fc.get('removed',[]))}, changed {len(fc.get('changed',[]))}")
        lines.append('')
    lines.append('## Blockers & Warnings')
    for label, items in (('Blockers', memo.get('blockers', [])), ('Warnings', memo.get('warnings', []))):
        lines.append(f"- {label}: " + ('; '.join(items[:3]) if items else 'None'))
    lines.append('')
    lines.append('## Next Actions')
    acts = memo.get('actions', [])
    if acts:
        for a in acts[:5]:
            lines.append(f"- {a}")
    else:
        lines.append('- No immediate actions recorded')
    (output_dir / 'client-summary.md').write_text('\n'.join(lines) + '\n', encoding='utf8')




def main() -> int:

    parser = argparse.ArgumentParser(description="Export Ship Sentinel audit JSON into spreadsheet deliverables.")
    parser.add_argument("input_json", help="Path to audit workspace JSON, app state JSON, or demo-state.js")
    parser.add_argument("--output-dir", help="Output directory for generated files")
    args = parser.parse_args()

    input_path = Path(args.input_json)
    if not input_path.is_absolute():
        input_path = (Path.cwd() / input_path).resolve()

    payload = normalize_payload(load_payload(input_path))
    audit_id = payload["meta"]["audit_id"]
    output_dir = Path(args.output_dir).resolve() if args.output_dir else (OUTPUT / audit_id).resolve()

    # Load previous snapshot (if any) before overwriting outputs
    prev_payload_path = output_dir / "normalized-payload.json"
    prev_metadata_path = output_dir / "package-metadata.json"
    prev_payload = json.loads(prev_payload_path.read_text(encoding="utf8")) if prev_payload_path.exists() else None
    prev_metadata = json.loads(prev_metadata_path.read_text(encoding="utf8")) if prev_metadata_path.exists() else None

    export_workbooks(payload, output_dir)
    export_release_memo(payload, output_dir)

    # Write normalized payload snapshot for diffing
    (output_dir / "normalized-payload.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2)  + "\n",
        encoding="utf8",
    )

    # Supporting files and metadata (includes file hashes)
    curr_metadata = write_supporting_files_and_metadata(payload, output_dir)

    # Compute package-to-package diff summary
    diff = {
        "audit_id": audit_id,
        "generated_at": curr_metadata.get("generated_at"),
        "previous_generated_at": (prev_metadata or {}).get("generated_at"),
        "file_changes": {},
        "payload_changes": compute_payload_changes(prev_payload, payload),
    }

    # File-level changes if previous metadata contains file_hashes
    prev_hashes = (prev_metadata or {}).get("file_hashes") or {}
    curr_hashes = curr_metadata.get("file_hashes", {})
    prev_files = set(prev_hashes.keys())
    curr_files = set(curr_hashes.keys())
    added_files = sorted(curr_files - prev_files)
    removed_files = sorted(prev_files - curr_files)
    changed_files = sorted([name for name in (curr_files & prev_files) if prev_hashes.get(name) != curr_hashes.get(name)])
    diff["file_changes"] = {
        "added": added_files,
        "removed": removed_files,
        "changed": changed_files,
    }

    (output_dir / "package-diff.json").write_text(
        json.dumps(diff, ensure_ascii=False, indent=2)  + "\n",
        encoding="utf8",
    )

    # Client-facing diff markdown
    write_package_diff_md(diff, output_dir)

    # Delivery history (audit trail)
    write_delivery_history(audit_id, output_dir, curr_metadata)

    # Compact client summary bundle
    write_client_summary(payload, diff, output_dir)



    print(f"Exported audit pack to {output_dir}")
    return 0
if __name__ == "__main__":
    raise SystemExit(main())



